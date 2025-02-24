# Import necessary modules from openpyxl for working with Excel files
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Define a class for extracting data from an Excel file
class DataExtract:
    def __init__(self):
        # Load the workbook from the specified Excel file
        self.wb = load_workbook("data/Sample-Data.xlsx")
        # Set the active worksheet (the one that opens by default)
        self.ws = self.wb.active
        # Initialize an empty list to store column headers
        self.columns = list()

    def get_columns(self):
        """
        Extract column headers from the 6th row and store them in self.columns.
        Assumes headers are in columns B to H (Excel columns 2 to 8).
        """
        for index in range(2, 9):
            # Get the value of the cell in row 6 and append it to the list of columns
            self.columns.append(self.ws[f"{get_column_letter(index)}6"].value)

    def get_data(self):
        """
        Extract data from rows 7 to 51 and convert it into a list of dictionaries.
        Each dictionary represents a row with column headers as keys.
        """
        data = list()   # List to store all extracted rows

        # Loop through each row from 7 to 51
        for row_index in range(7, 52):
            index = 0   # Tracks the position in the columns list
            set_dict = dict()   # Dictionary to hold cell data for the current row

            # Loop through columns B to H for each row
            for col_index in range(2, 9):
                # Assign the cell's value to the corresponding column header
                set_dict[self.columns[index]] = not self.ws[f"{get_column_letter(col_index)}{row_index}"].value
                index += 1  # Move to the next column header

            data.append(set_dict)   # Add the row data to the list

        return data     # Return the complete list of extracted data
