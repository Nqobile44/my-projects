from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import os
import time


def open_file():
    file_path = os.path.abspath("data/new_flights.xlsx")
    time.sleep(5)
    os.startfile(file_path)



class DataImport:
    def __init__(self, title: str, columns: list, data: dict):
        self.title = title
        self.columns = columns
        self.data = data

        self.wb = load_workbook(filename="data/new_flights.xlsx")
        self.ws = self.wb.active

    def extract(self):
        #setting up the title.

        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        thick_side_border = Border(
            right=Side(style="thick"),
            left=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        thick_top_border = Border(
            top=Side(style="thick")
        )

        #Setting up the columns border
        for column_index in range(1, 5):
            for index in range(9):
                self.ws[f"{get_column_letter(index+1)}{column_index+1}"].border = thick_border

        self.ws.merge_cells('A1:I3')
        self.ws["A1"] = self.title.title()
        self.ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['A1'].font = Font(size=30, bold=True)

        self.ws.merge_cells("A4:C4")
        self.ws.merge_cells("D4:F4")
        self.ws.merge_cells("G4:I4")

        headings = list(self.data.keys())

        # Setting up the columns.
        self.ws["A4"] = headings[0].title()
        self.ws["A4"].alignment = Alignment(horizontal="center")
        self.ws["A4"].font = Font(size=10, bold=True)
        self.ws["D4"] = headings[1].title()
        self.ws["D4"].alignment = Alignment(horizontal="center")
        self.ws["D4"].font = Font(size=10, bold=True)
        self.ws["G4"] = headings[2].title()
        self.ws["G4"].alignment = Alignment(horizontal="center")
        self.ws["G4"].font = Font(size=10, bold=True)

        main_columns = list(self.data["groceries"][0].keys())

        #settig up the main columns
        index = 1
        for value_index in range(3):
            self.ws[f"{get_column_letter(index)}5"] = main_columns[0].title()
            self.ws[f"{get_column_letter(index)}5"].font = Font(size=10, bold=True)
            self.ws[f"{get_column_letter(index+1)}5"] = main_columns[1].title()
            self.ws[f"{get_column_letter(index+1)}5"].font = Font(size=10, bold=True)
            self.ws[f"{get_column_letter(index+2)}5"] = main_columns[2].title()
            self.ws[f"{get_column_letter(index+2)}5"].font = Font(size=10, bold=True)
            index += 3

        #resizing the columns.
        for index in range(1, 11):
            self.ws.column_dimensions[get_column_letter(index)].width = 15

        # Setting up the border in value cells.
        for row_index in range(6, len(self.data["groceries"])+6):
            for column_index in range(1, 10):
                self.ws[f"{get_column_letter(column_index)}{row_index}"].border = thin_border

        # Setting up the separating borders
        column_index = 3
        for _ in range(3):
            for row_index in range(6, 8 + 6):
                self.ws[f"{get_column_letter(column_index)}{row_index}"].border = thick_side_border
            column_index += 3

        # Setting up the bottom thick border
        for col_index in range(1, 3*3+1):
            self.ws[f"{get_column_letter(col_index)}14"].border = thick_top_border

        # Setting up the fill in, in columns cell.
        for row_index in range(1, 6):
            for col_index in range(1, 10):
                self.ws[f"{get_column_letter(col_index)}{row_index}"].fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

        #setting up the 'groceries' value to cells.
        row_index = 6
        for column_index in self.data["groceries"]:
            self.ws[f"{get_column_letter(1)}{row_index}"] = column_index.get("item")
            self.ws[f"{get_column_letter(2)}{row_index}"] = column_index.get("quantity")
            self.ws[f"{get_column_letter(3)}{row_index}"] = column_index.get("unit")
            row_index += 1

        # setting up the 'personalCare' value to cells.
        row_index = 6
        for column_index in self.data["personalCare"]:
            self.ws[f"{get_column_letter(4)}{row_index}"] = column_index.get("item")
            self.ws[f"{get_column_letter(5)}{row_index}"] = column_index.get("quantity")
            self.ws[f"{get_column_letter(6)}{row_index}"] = column_index.get("unit")
            row_index += 1

        # setting up the 'householdItems' value to cells.
        row_index = 6
        for column_index in self.data["householdItems"]:
            self.ws[f"{get_column_letter(7)}{row_index}"] = column_index.get("item")
            self.ws[f"{get_column_letter(8)}{row_index}"] = column_index.get("quantity")
            self.ws[f"{get_column_letter(9)}{row_index}"] = column_index.get("unit")
            row_index += 1

    def save_data(self):
        self.wb.save(filename="data/new_flights.xlsx")
